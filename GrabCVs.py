#importing
import time
import requests
import threading
from bs4 import BeautifulSoup
import wget
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series

#time start
start = time.time()


#setting up page search array and base info
base_url = 'https://www.uwstout.edu/a-to-z-directory/'
directory_url = 'https://www.uwstout.edu'
cv_url = 'https://www.uwstout.edu/sites/default/files/cv/'
alphebet = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
page_num_text = "?page="
page_num=""
no_page = []

AllFacultyNames=[]
AllFacultyDepartments=[]
AllFacultyDirectories=[]
AllFacultyCVs=[]

#setting up subprocess to download links
def LinkReplacement(Link):
    if "%" in Link:
        Link=Link.replace("%20"," ")
        Link=Link.replace("%2C",",")
        Link=Link.replace("%21","!")
        Link=Link.replace("%23","#")
        Link=Link.replace("%24","$")
        Link=Link.replace("%26","&")
        Link=Link.replace("%28","(")
        Link=Link.replace("%29",")")
        Link=Link.replace("%2D","-")
        Link=Link.replace("%2E",".")
        Link=Link.replace("%5B","[")
        Link=Link.replace("%5D","]")
        Link=Link.replace("%5F","_")
        Link=Link.replace("%60","`")
        try:
            wget.download(Link)
        except:
            print("PAGE NOT FOUND: " + Link)
            with open('NotWorkingLinks.txt','a') as f:
                f.write(Link+"\n")
def DownloadLink(Link):
    try:
        wget.download(Link)
    except:
        LinkReplacement(Link)

#cycle through different pages to get names
finished = False
while (finished == False):
    #set current letter
    print("Starting Data Scraping")
    for letter in alphebet:
        page_num="0"
        
        #if next page continue
        No_Next = False
        while(No_Next == False):
            
            #finding if next page exists
            response = requests.get(base_url+letter+page_num_text+page_num)
            Button_Info_Search = BeautifulSoup(response.content, "html.parser")
            Next_Button = Button_Info_Search.find_all("a", {"class":"button pagination__next"})
            
            #if there isnt a next button, there is no next and you should move on to the next letter
            if Next_Button == no_page:
                No_Next = True
                
            #grabbing pages of directory
            response = requests.get(base_url+letter+page_num_text+page_num)
            PageInfo = BeautifulSoup(response.content, "html.parser")

            #grabbing each professor section
            faculty_info = PageInfo.find_all("div", {"class": "faculty-search__text"})
            
            #grabing info from faculty sections
            for faculty_section in faculty_info:
                #getting names and directory info (directory info for their pages is for the cvs later)
                name = faculty_section.find("h3", {"class": "faculty-search__name"})
                name = str(name)
                directory = name[name.find("<a"):name.find("</a>")]
                directory = directory[directory.find('"/')+1:directory.find('">')]
                AllFacultyDirectories.append(directory)
                name = name[name.find("<a"):name.find("</a>")]
                name = name[name.find(">")+1:]
                AllFacultyNames.append(name)

                #getting department if any
                sub_sections = faculty_section.find_all("div", {"class": "faculty-search__info-item"})
                department_found = False
                for item in sub_sections:
                    if department_found == False:
                        sub_section_item = item.find("span",{"class": "faculty-search__info-label"})
                        if "Department" in sub_section_item:
                            department_found = True
                            FoundDepartment = str(item.find("span",{"class": "faculty-search__info-data"}))
                            FoundDepartment = FoundDepartment[FoundDepartment.find('>')+1:FoundDepartment.find("</span>")]
                            AllFacultyDepartments.append(FoundDepartment)
                if department_found == False:
                    AllFacultyDepartments.append("N/A")
            #Canceling the loops when necesary
            if (No_Next == False):
                page_num = str(int(page_num)+1)
            else:
                print(letter + " Scraped")
            if (letter == "Z"):
                finished = True
#run through directories and find cv links
print("Fetching Staff Pages, downloading CVs may take a while")
count = 0
counter= int(len(AllFacultyDirectories)*.05)
percent=5
for link in AllFacultyDirectories:
    count = count+1
    if (count%counter==0):
        print("---"+str(percent)+"% complete")
        percent = percent+5
    response = requests.get(directory_url+link)
    StaffInfo = BeautifulSoup(response.content, "html.parser")
    additional = StaffInfo.find("div", {"class": "l-content--aside"})
    appended=False
    itemFound=False
    #try and find list items in the directory page, they have the possible link
    try:
        list_items = additional.find_all("li")
        for item in list_items:
            item = str(item)
            if 'href="/sites/default/files/cv/' in item:
                itemFound=True
                item=item[item.find("/cv/")+4:item.find('" target')]
                AllFacultyCVs.append("Y")
                appended=True
                #download link with multithreading for better speed
                true_link = cv_url+item
                download_thread = threading.Thread(target=DownloadLink, args=(true_link,))
                download_thread.start()
        if itemFound==False:
            AllFacultyCVs.append("N")
            appended=True  
    except:
        if appended==False:
            AllFacultyCVs.append("N")

#Excel Stuff
#set up workbook
print("Creating Excel Spreadsheet")
wb = Workbook()
ws = wb.active
#assign headers
ws.cell(1,1).value = "Last Name"
ws.cell(1,2).value = "First Name"
ws.cell(1,3).value = "Department"
ws.cell(1,4).value = "If CV"
for x in range (len(AllFacultyNames)):
    #go through all faculty in directory. if there is a first and last name seperate them, if not make last name only name
    if "," in AllFacultyNames[x]:
        new_name = AllFacultyNames[x].split(",")
        ws.cell(x+2,1).value = new_name[0]
        ws.cell(x+2,2).value = new_name[1]
    else:
        ws.cell(x+2,1).value = AllFacultyNames[x]
        ws.cell(x+2,2).value = "N/A"
    #add department
    ws.cell(x+2,3).value = AllFacultyDepartments[x]
    #add if CV
    ws.cell(x+2,4).value = AllFacultyCVs[x]
#save workbook
wb.save('CVSHEET.xlsx')

####end time
end = time.time()
#state time spent
print("Time Spent: "+str(end-start)+" seconds")

#
#link format
#https://www.uwstout.edu/a-to-z-directory/A?page=0
#time spent, 6:30
    
